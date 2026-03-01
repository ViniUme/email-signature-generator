#!/usr/bin/env python3
"""Creates sample Excel files in the data/ folder for testing."""

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '--quiet'])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

DATA_DIR = Path(__file__).parent / 'data'
DATA_DIR.mkdir(exist_ok=True)


def create_excel(filename: str, employees: list, sheet_name: str = "Employees"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = ['Name', 'Position', 'Phone', 'Email', 'Department']
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row, emp in enumerate(employees, 2):
        for col, val in enumerate(emp, 1):
            ws.cell(row=row, column=col, value=val)

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 34
    ws.column_dimensions['E'].width = 20

    path = DATA_DIR / filename
    wb.save(str(path))
    print(f"Created: {path}")


# ── Sample file 1: Technology department ──────────────────────────────────────
create_excel("employees_technology.xlsx", [
    ['Ana Silva',       'Project Manager',       '(11) 98765-4321', 'ana.silva@company.com',       'Technology'],
    ['Bruno Santos',    'Systems Analyst',        '(11) 91234-5678', 'bruno.santos@company.com',    'Technology'],
    ['Daniel Pereira',  'Senior Developer',       '(11) 94567-8901', 'daniel.pereira@company.com',  'Technology'],
    ['Felipe Martins',  'UX/UI Designer',         '(11) 93456-7890', 'felipe.martins@company.com',  'Technology'],
    ['Laura Rocha',     'QA Engineer',            '(11) 92345-6789', 'laura.rocha@company.com',     'Technology'],
])

# ── Sample file 2: Full company ───────────────────────────────────────────────
create_excel("employees_all.xlsx", [
    ['Hugo Ferreira',   'CEO',                    '(11) 99999-0000', 'hugo.ferreira@company.com',   'Executive'],
    ['Carla Oliveira',  'Commercial Director',    '(11) 99876-5432', 'carla.oliveira@company.com',  'Commercial'],
    ['Elena Costa',     'HR Coordinator',         '(11) 98901-2345', 'elena.costa@company.com',     'Human Resources'],
    ['Gabriela Lima',   'Financial Analyst',      '(11) 97890-1234', 'gabriela.lima@company.com',   'Finance'],
    ['Ana Silva',       'Project Manager',        '(11) 98765-4321', 'ana.silva@company.com',       'Technology'],
    ['Bruno Santos',    'Systems Analyst',        '(11) 91234-5678', 'bruno.santos@company.com',    'Technology'],
    ['Daniel Pereira',  'Senior Developer',       '(11) 94567-8901', 'daniel.pereira@company.com',  'Technology'],
    ['Felipe Martins',  'UX/UI Designer',         '(11) 93456-7890', 'felipe.martins@company.com',  'Technology'],
])

print("\nSample Excel files created in the data/ folder.")
print("Run the generator with: python email_signature_generator.py")
